import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Maximum characters to store from full-text extraction (~1 MB of text)
_MAX_FULLTEXT_CHARS = 1_000_000


class DmsFile(models.Model):
    _inherit = "dms.file"

    def _check_access_dms_record(self, operation):
        """Fix: include archived records in access check to allow unarchive."""
        if any(self._ids) and not self.env.su:
            Rule = self.env["ir.rule"]
            domain = Rule._compute_domain(self._name, operation)
            items = self.with_context(active_test=False).search(domain)
            if any(x_id not in items.ids for x_id in self.ids):
                raise Rule._make_access_error(operation, (self - items))

    # ------------------------------------------------------------------
    # GED fields (v1 — existing)
    # ------------------------------------------------------------------
    document_type_id = fields.Many2one(
        "isic.document.type",
        string="Type de document",
        tracking=True,
        ondelete="set null",
    )
    ged_state = fields.Selection(
        [
            ("draft", "Brouillon"),
            ("validated", "Validé"),
            ("archived", "Classé"),
        ],
        string="État GED",
        default="draft",
        tracking=True,
    )
    reference = fields.Char(
        string="Référence",
        index=True,
        copy=False,
        tracking=True,
    )
    date_document = fields.Date(
        string="Date du document",
        tracking=True,
    )
    annee_academique_id = fields.Many2one(
        "isic.annee.academique",
        string="Année académique",
        tracking=True,
        ondelete="set null",
    )
    valideur_id = fields.Many2one(
        "res.users",
        string="Validé par",
        readonly=True,
        ondelete="set null",
    )
    date_validation = fields.Datetime(
        string="Date de validation",
        readonly=True,
    )

    # ------------------------------------------------------------------
    # V2 — Versionning
    # ------------------------------------------------------------------
    version_ids = fields.One2many(
        "isic.document.version",
        "file_id",
        string="Historique des versions",
    )
    version_count = fields.Integer(
        string="Nb versions",
        compute="_compute_version_count",
        store=True,
    )
    current_version = fields.Integer(
        string="Version courante",
        readonly=True,
        default=0,
    )

    @api.depends("version_ids")
    def _compute_version_count(self):
        for rec in self:
            rec.version_count = len(rec.version_ids)

    # ------------------------------------------------------------------
    # V2 — Full-text search
    # ------------------------------------------------------------------
    fulltext_content = fields.Text(
        string="Contenu textuel",
        readonly=True,
    )
    fulltext_indexed = fields.Boolean(
        string="Indexé",
        readonly=True,
        default=False,
    )
    fulltext_error = fields.Char(
        string="Erreur d'indexation",
        readonly=True,
    )

    # ------------------------------------------------------------------
    # V2 — Preview
    # ------------------------------------------------------------------
    preview_type = fields.Selection(
        [
            ("pdf", "PDF"),
            ("image", "Image"),
            ("none", "Aucune"),
        ],
        string="Type de prévisualisation",
        compute="_compute_preview_type",
        store=True,
    )

    # ------------------------------------------------------------------
    # V2 — Classification automatique
    # ------------------------------------------------------------------
    auto_classified = fields.Boolean(
        string="Classifié automatiquement",
        readonly=True,
        default=False,
    )

    # ==================================================================
    # Compute methods
    # ==================================================================

    @api.depends("mimetype", "extension")
    def _compute_preview_type(self):
        image_mimes = {"image/png", "image/jpeg", "image/jpg", "image/gif", "image/svg+xml", "image/webp"}
        image_exts = {"png", "jpg", "jpeg", "gif", "svg", "webp"}
        for rec in self:
            mime = rec.mimetype or ""
            ext = (rec.extension or "").lower().lstrip(".")
            if mime == "application/pdf" or ext == "pdf":
                rec.preview_type = "pdf"
            elif mime in image_mimes or ext in image_exts:
                rec.preview_type = "image"
            else:
                rec.preview_type = "none"

    # ==================================================================
    # Preview action (open in new tab)
    # ==================================================================

    def action_preview(self):
        """Open preview in a new browser tab via URL redirect."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_url",
            "url": f"/isic_ged/preview/{self.id}",
            "target": "new",
        }

    # ==================================================================
    # Thumbnail refresh
    # ==================================================================

    def _recompute_thumbnail(self):
        """Force thumbnail update after content change.

        The base DMS compute depends on 'content' (a non-stored field),
        which doesn't reliably trigger recomputation of the stored image_1920.
        """
        self.ensure_one()
        try:
            from PIL import Image as PILImage

            mime = self.mimetype or ""
            supported = {*PILImage.MIME.values(), "image/svg+xml"} - {"application/pdf"}
            if mime in supported and self.content:
                self.image_1920 = self.content
            else:
                # Non-image file (PDF, DOCX, etc.) — clear stale thumbnail
                if self.image_1920:
                    self.image_1920 = False
        except Exception:
            pass

    # ==================================================================
    # GED workflow (v1)
    # ==================================================================

    def action_validate(self):
        for rec in self:
            if rec.ged_state != "draft":
                raise UserError(_("Seul un document en brouillon peut être validé."))
            rec.write(
                {
                    "ged_state": "validated",
                    "valideur_id": self.env.uid,
                    "date_validation": fields.Datetime.now(),
                }
            )

    def action_archive_ged(self):
        for rec in self:
            if rec.document_type_id.validation_required and rec.ged_state != "validated":
                raise UserError(_("Ce type de document nécessite une validation avant classement."))
            rec.write({"ged_state": "archived"})

    def action_reset_draft(self):
        if not self.env.user.has_group("isic_base.group_isic_direction"):
            raise UserError(_("Seule la direction peut remettre un document en brouillon."))
        self.write({"ged_state": "draft", "valideur_id": False, "date_validation": False})

    @api.onchange("document_type_id")
    def _onchange_document_type_id(self):
        if self.document_type_id and not self.annee_academique_id:
            self.annee_academique_id = self.env["isic.annee.academique"]._get_current()

    # ==================================================================
    # Versionning
    # ==================================================================

    def _create_version(self, comment=""):
        """Snapshot the current content as a new version before overwrite."""
        for rec in self:
            if not rec.content:
                continue
            next_version = rec.current_version + 1
            self.env["isic.document.version"].sudo().create(
                {
                    "file_id": rec.id,
                    "version_number": next_version,
                    "content": rec.content,
                    "checksum": rec.checksum,
                    "size": rec.size,
                    "mimetype": rec.mimetype,
                    "name": rec.name,
                    "author_id": self.env.uid,
                    "comment": comment,
                }
            )
            # Update current_version via SQL to avoid re-triggering write()
            self.env.cr.execute(
                "UPDATE dms_file SET current_version = %s WHERE id = %s",
                (next_version, rec.id),
            )
            rec.invalidate_recordset(["current_version"])

            # Purge old versions beyond limit
            max_versions = int(self.env["ir.config_parameter"].sudo().get_param("isic_ged.max_versions", default=50))
            versions = rec.version_ids.sorted("version_number", reverse=True)
            if len(versions) > max_versions:
                versions[max_versions:].unlink()

    def action_restore_version(self):
        """Restore a previous version. Called from version list button.

        Expects the version id in the context key 'active_id'.
        """
        self.ensure_one()
        version_id = self.env.context.get("restore_version_id")
        if not version_id:
            raise UserError(_("Aucune version sélectionnée pour la restauration."))

        version = self.env["isic.document.version"].browse(version_id)
        if not version.exists() or version.file_id != self:
            raise UserError(_("Version invalide."))

        # Save current state as a new version before restoring
        self._create_version(comment=_("Sauvegarde avant restauration de v%s", version.version_number))

        # Restore the old content (bypass versioning via context flag)
        self.with_context(_isic_skip_version=True).write({"content": version.content, "name": version.name})
        self.message_post(body=_("Document restauré à la version %s.", version.version_number))

    # ==================================================================
    # Full-text extraction & indexing
    # ==================================================================

    def _extract_text_content(self):
        """Extract text from file content based on mimetype.

        Supported formats: PDF (pypdf), DOCX (python-docx), XLSX (openpyxl), plain text.

        Uses direct SQL to avoid re-entering the write() override when called
        from create() or write().
        """
        # Flush any pending ORM writes on fulltext fields before raw SQL updates
        self.flush_recordset(["fulltext_content", "fulltext_indexed", "fulltext_error"])
        for rec in self:
            if not rec.content:
                self.env.cr.execute(
                    "UPDATE dms_file SET fulltext_content = '', fulltext_indexed = FALSE, fulltext_error = '' WHERE id = %s",
                    (rec.id,),
                )
                rec.invalidate_recordset(["fulltext_content", "fulltext_indexed", "fulltext_error"])
                continue

            try:
                binary = base64.b64decode(rec.content)
                text = ""
                mime = rec.mimetype or ""

                if mime == "application/pdf":
                    text = rec._extract_pdf(binary)
                elif mime in (
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    "application/msword",
                ):
                    text = rec._extract_docx(binary)
                elif mime in (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                ):
                    text = rec._extract_xlsx(binary)
                elif mime.startswith("text/"):
                    text = binary.decode("utf-8", errors="replace")

                if len(text) > _MAX_FULLTEXT_CHARS:
                    text = text[:_MAX_FULLTEXT_CHARS]

                self.env.cr.execute(
                    "UPDATE dms_file SET fulltext_content = %s, fulltext_indexed = %s, fulltext_error = '' WHERE id = %s",
                    (text, bool(text), rec.id),
                )
                rec.invalidate_recordset(["fulltext_content", "fulltext_indexed", "fulltext_error"])
            except Exception as e:
                _logger.warning("Full-text extraction failed for file %s: %s", rec.id, e)
                self.env.cr.execute(
                    "UPDATE dms_file SET fulltext_content = '', fulltext_indexed = FALSE, fulltext_error = %s WHERE id = %s",
                    (str(e)[:200], rec.id),
                )
                rec.invalidate_recordset(["fulltext_content", "fulltext_indexed", "fulltext_error"])

    @staticmethod
    def _extract_pdf(binary):
        """Extract text from PDF binary using pypdf."""
        try:
            from pypdf import PdfReader
        except ImportError:
            _logger.info("pypdf not installed, skipping PDF text extraction")
            return ""

        reader = PdfReader(io.BytesIO(binary))
        parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                parts.append(text)
        return "\n".join(parts)

    @staticmethod
    def _extract_docx(binary):
        """Extract text from DOCX binary using python-docx."""
        try:
            from docx import Document
        except ImportError:
            _logger.info("python-docx not installed, skipping DOCX text extraction")
            return ""

        doc = Document(io.BytesIO(binary))
        return "\n".join(para.text for para in doc.paragraphs if para.text)

    @staticmethod
    def _extract_xlsx(binary):
        """Extract text from XLSX binary using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            _logger.info("openpyxl not installed, skipping XLSX text extraction")
            return ""

        wb = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" ".join(cells))
        wb.close()
        return "\n".join(parts)

    def _update_fulltext_index(self):
        """Update the PostgreSQL tsvector column for full-text search.

        The tsvector column is created by post_init_hook (first install).
        On module update, if the column is missing, create it on the fly.
        """
        self.env.cr.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name = 'dms_file' AND column_name = 'fulltext_tsvector'"
        )
        if not self.env.cr.fetchone():
            self.env.cr.execute("ALTER TABLE dms_file ADD COLUMN fulltext_tsvector tsvector")
            self.env.cr.execute(
                "CREATE INDEX IF NOT EXISTS idx_dms_file_fulltext ON dms_file USING gin(fulltext_tsvector)"
            )
            _logger.info("Created fulltext_tsvector column and GIN index on dms_file")
        for rec in self:
            if rec.fulltext_content:
                self.env.cr.execute(
                    """
                    UPDATE dms_file
                    SET fulltext_tsvector = to_tsvector('french', %(content)s)
                    WHERE id = %(id)s
                    """,
                    {"content": rec.fulltext_content, "id": rec.id},
                )
            else:
                self.env.cr.execute(
                    "UPDATE dms_file SET fulltext_tsvector = NULL WHERE id = %s",
                    (rec.id,),
                )

    @api.model
    def search_fulltext(self, query, limit=80):
        """Full-text search across indexed documents using PostgreSQL tsvector.

        :param query: Search terms (French language stemming applied)
        :param limit: Max results
        :return: dms.file recordset matching the query (access-rule filtered)
        """
        if not query or not query.strip():
            return self.browse()

        self.env.cr.execute(
            """
            SELECT id, ts_rank(fulltext_tsvector, plainto_tsquery('french', %(query)s)) AS rank
            FROM dms_file
            WHERE fulltext_tsvector @@ plainto_tsquery('french', %(query)s)
            ORDER BY rank DESC
            LIMIT %(limit)s
            """,
            {"query": query.strip(), "limit": limit},
        )
        results = self.env.cr.fetchall()
        if not results:
            return self.browse()

        ids = [r[0] for r in results]
        return self.search([("id", "in", ids)])

    def action_reindex_fulltext(self):
        """Server action: reindex full-text content for selected files."""
        self._extract_text_content()
        self._update_fulltext_index()

    # ==================================================================
    # Classification automatique
    # ==================================================================

    def _auto_classify(self):
        """Apply classification rules on files that haven't been manually classified."""
        rules = self.env["isic.document.classification.rule"].search([])
        if not rules:
            return

        for rec in self:
            # Skip if already manually classified
            if rec.document_type_id and not rec.auto_classified:
                continue

            for rule in rules:
                if rule._match(rec):
                    vals = {"auto_classified": True}
                    if rule.document_type_id:
                        vals["document_type_id"] = rule.document_type_id.id
                    if rule.tag_ids:
                        vals["tag_ids"] = [(4, tid) for tid in rule.tag_ids.ids]
                    rec.with_context(_isic_skip_version=True).write(vals)
                    break

    # ==================================================================
    # CRUD overrides
    # ==================================================================

    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        # Auto-classify new files
        records._auto_classify()
        # Extract full-text content for new files
        records._extract_text_content()
        records._update_fulltext_index()
        return records

    # Fields protected when document is validated/archived
    _PROTECTED_FIELDS = {"content", "name"}

    def write(self, vals):
        # ------ Protection: validated/archived documents ------
        if not self.env.context.get("_isic_skip_version"):
            protected = self._PROTECTED_FIELDS & set(vals)
            if protected:
                for rec in self:
                    if rec.ged_state in ("validated", "archived"):
                        raise UserError(
                            _(
                                "Le document « %s » est %s. Remettez-le en brouillon avant de le modifier.",
                                rec.name,
                                dict(rec._fields["ged_state"].selection).get(rec.ged_state, rec.ged_state),
                            )
                        )

        # ------ Protection: DMS lock ------
        if not self.env.su:
            protected_by_lock = self._PROTECTED_FIELDS & set(vals)
            if protected_by_lock:
                for rec in self:
                    if rec.locked_by and rec.locked_by.id != self.env.uid:
                        raise UserError(_("Le document « %s » est verrouillé par %s.", rec.name, rec.locked_by.name))

        # Versionning: snapshot before content change
        if "content" in vals and not self.env.context.get("_isic_skip_version"):
            self._create_version()

        res = super().write(vals)

        # Force thumbnail refresh when content changes
        if "content" in vals:
            for rec in self:
                rec._recompute_thumbnail()

        # Re-classify if name or directory changed
        if ("name" in vals or "directory_id" in vals) and not self.env.context.get("_isic_skip_version"):
            self._auto_classify()

        # Re-index if content changed
        if "content" in vals:
            self._extract_text_content()
            self._update_fulltext_index()

        return res
